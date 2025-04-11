import os
import re
import json
import openai
import logging
import tkinter
import traceback
import threading
import pandas as pd
import customtkinter
from typing import Tuple
from datetime import date
from tkcalendar import DateEntry
import tkinter.messagebox as msg
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bot import WebAccess, SharePoint, PDFProcessor

customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")


def HandleException(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            msg.showerror("Lỗi", f"{e} \n {traceback.format_exc()}")

    return wrapper


# ========== Frames ==========
class AccessCheck(customtkinter.CTkFrame):
    def __init__(self, parent, logger: logging.Logger):
        super().__init__(parent)
        self.grid_rowconfigure(3, weight=1)  # Log box ở dòng 3 sẽ mở rộng
        self.grid_columnconfigure(0, weight=1)
        self.logger = logger
        # Title
        self.title_label = customtkinter.CTkLabel(
            self, text="AccessCheck", font=("Arial", 20)
        )
        self.title_label.grid(row=0, column=0, pady=(20, 10))
        # DateEntry Frame
        datetime_frame = customtkinter.CTkFrame(self)
        datetime_frame.grid(row=1, column=0, pady=20)

        from_date = tkinter.Frame(datetime_frame, width=200, height=50)
        from_date.grid(row=0, column=0, padx=20, pady=10)

        to_date = tkinter.Frame(datetime_frame, width=200, height=50)
        to_date.grid(row=0, column=1, padx=20, pady=10)

        self.date_start = DateEntry(
            from_date, width=15, foreground="white", date_pattern="yyyy-mm-dd"
        )
        self.date_start.pack()

        self.date_end = DateEntry(
            to_date, width=15, foreground="white", date_pattern="yyyy-mm-dd"
        )
        self.date_end.pack()

        # Start button
        self.start_btn = customtkinter.CTkButton(
            self,
            text="RUN",
            hover=True,
            state="normal",
            command=self.active,
        )
        self.start_btn.grid(row=2, column=0, pady=(5, 5))
        # Log box
        self.log_box = customtkinter.CTkTextbox(self)
        self.log_box.configure(state="disabled")
        self.log_box.grid(
            row=3,
            column=0,
            sticky="nsew",
            padx=20,
            pady=10,
        )

    @HandleException
    def __update_log(self):
        initial_line_count = 0
        if os.path.exists("bot.log"):
            with open("bot.log", "r", encoding="utf-8") as f:
                initial_line_count = sum(1 for _ in f)

        while self.start_btn.cget("state") == "disabled":
            if os.path.exists("bot.log"):
                with open("bot.log", "r", encoding="utf-8") as f:
                    all_lines = f.readlines()
                    new_lines = all_lines[initial_line_count:]
                    if new_lines:
                        self.log_box.configure(state="normal")
                        self.log_box.insert("end", "".join(new_lines))
                        self.log_box.see("end")
                        self.log_box.configure(state="disabled")
                    initial_line_count = len(all_lines)
            else:
                self.log_box.configure(state="disabled")

    @HandleException
    def access_check(
        self,
        start_date: date,
        to_date: date,
    ):
        self.start_btn.configure(state="disabled")

        threading.Thread(
            target=self.__update_log,
            daemon=True,
        ).start()

        # Process Excel File
        excelPath = "Builder Sort.xlsm"
        if not os.path.exists(excelPath):
            self.start_btn.configure(state="normal")
            raise Exception("Không tìm thấy Build Sort.xlsm")
        # ------- Clean Sheet ------- #
        for sheet_name in ["Osaka", "Yokohama"]:
            wb = load_workbook(excelPath, keep_vba=True)
            if sheet_name in wb.sheetnames:
                std = wb[sheet_name]
                wb.remove(std)
                wb.create_sheet(title=sheet_name)
                wb.save(excelPath)
        # ------- Load Data from WebAccess ------- #
        wb = load_workbook(excelPath, keep_vba=True)
        BuilderSort = pd.read_excel(
            io=excelPath,
            sheet_name="BuilderList",
            dtype={
                "正しいビルダーコード": str,
            },
        )
        for _, row in BuilderSort.iterrows():
            df = WebAccess(
                username="2909",
                password="159753",
            ).get_information(
                ビルダー名=row["正しいビルダーコード"],
                確定納品日=[
                    start_date.strftime("%Y/%m/%d"),
                    to_date.strftime("%Y/%m/%d"),
                ],
            )
            if df is None or df.empty:
                continue
            target_sheet = (
                "Yokohama" if row["ビルダー名"] == "タマホーム㈱" else "Osaka"
            )
            ws = wb[target_sheet]
            start_row = ws.max_row + 1 if ws.max_row > 1 else 1
            # Append data df -> sheet Excel
            for r_idx, r in enumerate(
                dataframe_to_rows(df, index=False, header=(ws.max_row == 1))
            ):
                for c_idx, value in enumerate(r, 1):
                    ws.cell(row=start_row + r_idx, column=c_idx, value=value)
        wb.save(excelPath)
        # ------- Download List PDF File in SharePoint ------- #
        Yokohama = pd.read_excel(
            io=excelPath,
            sheet_name="Yokohama",
        )
        SHAREPOINT_DOWNLOAD_PATH = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "SharePoint Downloads"
        )
        SP = SharePoint(
            url="https://nskkogyo.sharepoint.com/",
            username="vietnamrpa@nskkogyo.onmicrosoft.com",
            password="Robot159753",
            download_directory=SHAREPOINT_DOWNLOAD_PATH,
        )
        for url in Yokohama["資料リンク"].to_list():
            # Download PDF
            success, _ = SP.download_file(
                site_url=url,
                file_pattern="割付図・エクセル/.*.pdf$",
            )
            if not success:
                success, _ = SP.download_file(
                    site_url=url,
                    file_pattern="割付図。エクセル/.*.pdf$",
                )
            # Download Excel
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.support.wait import WebDriverWait
            from selenium.webdriver.common.by import By
            try:
                mitsumorisho = WebDriverWait(SP.browser,60).wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'見積')]"))
                )
                if mitsumorisho.text.endswith(('xlsx','xls', 'xlsm')):
                    print("mitsumorisho excel file found.")
                else:
                    print("mitsumorisho folder found.")
            except: 
                pass

        del SP
        # ------- Extract Data from PDF  ------- #
        pdfProcessor = PDFProcessor(
            poppler_path=r"D:\VanNgocNhatHuy\RPA\AccessCheck\bin",
        )
        for file in os.listdir(SHAREPOINT_DOWNLOAD_PATH):
            pdf_path = os.path.join(SHAREPOINT_DOWNLOAD_PATH, file)
            full_text = pdfProcessor.extract_text_from_pdf(pdf_path)
            if full_text:
                data = self.__query_openai_for_data(full_text)
                data = json.loads(data)
                print(data)
        # ------- Merge Data  ------- #

        # ------- Return ------- #
        msg.showinfo("Thông Báo", "AccessCheck hoàn thành")
        self.start_btn.configure(state="normal")

    @HandleException
    def __query_openai_for_data(self, text: str):
        prompt_text = f"""
            Analyze the text and extract structured information with details for each floor (if multiple floors are mentioned). Give me the summarized values of 'Builder name', 'Stud sink direction', 'Floor number', 'Order number', 'Order name','Floor area 1', 'Floor area 2', 'Floor area 3', 'Loft', 'Penthouse area'. For floor number, return an array of floor numbers mentioned in the document. If there is only one page in the pdf, return ["1"].
            The values for floor areas should be mapped as follows:
            - 'Floor area 1': Area of 1st floor (numeric value only, without m² or ㎡)
            - 'Floor area 2': Area of 2nd floor (numeric value only, without m² or ㎡)
            - 'Floor area 3': Area of 3rd floor (numeric value only, without m² or ㎡)
            - 'Loft': Area marked as loft space (ロフト) or attic storage (小屋裏収納) (numeric value only)
            - 'Penthouse area': Area specifically marked as penthouse (numeric value only)
            Only return the numeric part of the floor area (e.g., "75.5" instead of "75.5m²"). The output should be in JSON format containing only these fields and their associated values:
            - Builder name (ビルダー名)
            - Stud sink direction (スタッド流し⽅向) (Format: '@number')
            - Floor number (Array of floor numbers, e.g., ["1"] for single floor, ["1", "2"] for two floors)
            - Order number (【案件No】) (Format: '6 digits')
            - Order name (【案件名】) (Format: name)
            - Floor area (【⾯積】) (Format: numeric values only, without units)
            - Penthouse area (Format: numeric value only)
            Text provided:
            {text}
            """
        try:
            response = openai.ChatCompletion.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "system",
                        "content": "You are a skilled assistant trained in extracting precise information from the pdf. Always respond with valid JSON, including all specified fields even if the value is empty or not found.",
                    },
                    {"role": "user", "content": prompt_text},
                ],
            )
            content = response["choices"][0]["message"]["content"]
            content = re.sub(r"^```json\s*|```\s*$", "", content, flags=re.MULTILINE)
            return content.strip()
        except Exception as e:
            self.logger.error(e)
            return None

    @HandleException
    def active(self):
        start_date = self.date_start.get_date()
        to_date = self.date_end.get_date()
        if to_date < start_date:
            raise Exception(
                "Vui lòng chọn ngày kết thúc lớn hơn hoặc bằng ngày bắt đầu."
            )
        threading.Thread(
            target=self.access_check,
            args=(start_date, to_date),
            daemon=True,
        ).start()


# ========== Main App ==========
class App(customtkinter.CTk):
    def __init__(
        self,
        title: str,
        geometry: str = None,
        resizable: Tuple[bool, bool] = (False, False),
        icon: str = None,
        logger_name: str = __name__,
    ):
        super().__init__()
        self.title(title)
        self.geometry(geometry)
        self.resizable(resizable[0], resizable[1])
        self.iconbitmap(icon)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.logger = logging.getLogger(logger_name)
        # Sidebar
        self.sidebar_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, sticky="ns")

        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame,
            text="TOOLS",
            font=customtkinter.CTkFont(size=20, weight="bold"),
        )
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.WebAccess = customtkinter.CTkButton(
            self.sidebar_frame,
            text="Access Check",
            command=lambda: self.SwitchTab("AccessCheck"),
        )
        self.WebAccess.grid(row=1, column=0, padx=20, pady=10)

        self.WebAccess = customtkinter.CTkButton(
            self.sidebar_frame,
            text="Mail Dealer",
            command=lambda: self.SwitchTab("MailDealer"),
        )
        self.WebAccess.grid(row=2, column=0, padx=20, pady=10)
        self.sidebar_frame.grid_rowconfigure(3, weight=1)
        self.appearance_mode_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Appearance Mode", anchor="w"
        )
        self.appearance_mode_label.grid(row=4, column=0, padx=20, pady=(10, 0))

        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(
            self.sidebar_frame,
            values=["Light", "Dark", "System"],
            command=self.ChangeAppearanceMode,
        )
        self.appearance_mode_optionemenu.grid(row=4, column=0, padx=20, pady=(10, 10))

        # Content Area
        self.content_frame = customtkinter.CTkFrame(self)
        self.content_frame.grid(row=0, column=1, sticky="nsew")  # Giãn hết phần còn lại

        self.content_frame.grid_rowconfigure(0, weight=1)
        self.content_frame.grid_columnconfigure(0, weight=1)

        # Khởi tạo các view
        self.views = {
            "AccessCheck": AccessCheck(parent=self.content_frame, logger=self.logger),
        }
        # Hiện view mặc định
        self.current_view = None
        self.SwitchTab("AccessCheck")

    @HandleException
    def SwitchTab(self, view_name):
        if self.current_view:
            self.views[self.current_view].grid_forget()
        if view_name in self.views.keys():
            self.views[view_name].grid(row=0, column=0, sticky="nsew")
            self.current_view = view_name
        else:
            msg.showinfo("Thông Báo","Chưa Hỗ Trợ")
            return

    @HandleException
    def ChangeAppearanceMode(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)